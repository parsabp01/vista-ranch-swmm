#!/usr/bin/env python3
"""Core SWMM automation pipeline stages."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from collections import defaultdict

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parents[1]


@dataclass
class StageResult:
    stage: str
    status: str
    message: str
    produced_files: List[str]


class PipelineContext:
    def __init__(self) -> None:
        self.extraction_cfg = self._load_yaml(ROOT / "configs" / "extraction.yaml")
        self.defaults_cfg = self._load_yaml(ROOT / "configs" / "defaults.yaml")

    @staticmethod
    def _load_yaml(path: Path) -> dict:
        with path.open("r", encoding="utf-8") as f:
            return yaml.safe_load(f) or {}


def now_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ensure_dirs() -> None:
    for rel in ["data/interim", "data/processed", "outputs/logs", "outputs/qa", "models"]:
        (ROOT / rel).mkdir(parents=True, exist_ok=True)


def required_paths(ctx: PipelineContext) -> Dict[str, List[Path] | Path]:
    req = ctx.extraction_cfg["required_inputs"]
    return {
        "workbook": ROOT / req["workbook"],
        "pdfs": [ROOT / p for p in req["pdfs"]],
        "planning_doc": ROOT / req["planning_doc"],
    }


def run_preflight(ctx: PipelineContext) -> StageResult:
    ensure_dirs()
    paths = required_paths(ctx)
    files_meta: List[dict] = []

    all_required: List[Tuple[str, Path]] = [
        ("workbook", paths["workbook"]),
        ("planning_doc", paths["planning_doc"]),
    ] + [("pdf", p) for p in paths["pdfs"]]

    blocked = False
    notes: List[str] = []
    for kind, p in all_required:
        exists = p.exists()
        if not exists:
            blocked = True
            notes.append(f"Missing required {kind}: {p.relative_to(ROOT)}")
        files_meta.append(
            {
                "kind": kind,
                "path": str(p.relative_to(ROOT)),
                "exists": exists,
                "size_bytes": p.stat().st_size if exists else None,
                "sha256": sha256_file(p) if exists else None,
                "modified_time_utc": datetime.fromtimestamp(p.stat().st_mtime, tz=timezone.utc).isoformat()
                if exists
                else None,
            }
        )

    status = "blocked" if blocked else "ready"
    manifest = {
        "generated_at_utc": now_utc(),
        "required_inputs": {
            "workbook": str(paths["workbook"].relative_to(ROOT)),
            "pdfs": [str(p.relative_to(ROOT)) for p in paths["pdfs"]],
            "planning_doc": str(paths["planning_doc"].relative_to(ROOT)),
        },
        "files": files_meta,
        "status": status,
        "notes": notes,
    }

    manifest_path = ROOT / "outputs/logs/intake_manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    return StageResult("preflight", status, "Required input check complete", [str(manifest_path.relative_to(ROOT))])


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower().replace("\n", " ")


def row_to_dict(headers: List[str], row: Iterable[object]) -> dict:
    return {headers[i]: row[i] for i in range(min(len(headers), len(row))) if headers[i]}


def classify_record(record: dict, keywords: Dict[str, List[str]]) -> str | None:
    blob = " ".join(str(v).lower() for v in record.values() if v is not None)
    scores = {group: sum(1 for kw in kws if kw in blob) for group, kws in keywords.items()}
    best_group, best_score = max(scores.items(), key=lambda item: item[1])
    return best_group if best_score > 0 else None


def write_csv(path: Path, rows: List[dict]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fields: List[str] = sorted({k for row in rows for k in row.keys()})
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def _to_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        val = float(value)
        return val if math.isfinite(val) else None
    text = str(value).strip()
    if text == "":
        return None
    match = re.search(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None




def _extract_hydraulic_links_from_sheet(all_rows: List[tuple], sheet_name: str) -> List[dict]:
    links: List[dict] = []
    header_idx = None
    for i, row in enumerate(all_rows):
        row_txt = " ".join(str(v).lower() for v in row if v is not None)
        if row_txt.count("stream jct") >= 2:
            header_idx = i
            break
    if header_idx is None:
        return links

    # table starts after units/header rows; robustly scan following rows
    for r in range(header_idx + 1, len(all_rows)):
        row = all_rows[r]
        vals = [v for v in row]
        if len(vals) < 8:
            continue
        ds = vals[1] if len(vals) > 1 else None
        us = vals[2] if len(vals) > 2 else None
        dia = _to_float(vals[6] if len(vals) > 6 else None)
        q = _to_float(vals[7] if len(vals) > 7 else None)
        length = _to_float(vals[8] if len(vals) > 8 else None)
        slope = _to_float(vals[11] if len(vals) > 11 else None)
        hl = _to_float(vals[12] if len(vals) > 12 else None)
        jtype = str(vals[14]).strip() if len(vals) > 14 and vals[14] is not None else ""

        if us is None or ds is None:
            continue
        us_str = str(us).strip()
        ds_str = str(ds).strip()
        if not us_str or not ds_str:
            continue
        if _to_float(us_str) is None or _to_float(ds_str) is None:
            continue

        links.append(
            {
                "pipe_id": f"{us_str}-{ds_str}",
                "upstream_node": us_str,
                "downstream_node": ds_str,
                "dia": dia,
                "q": q,
                "length": length,
                "slope": slope,
                "headloss": hl,
                "junction_type": jtype,
                "source_sheet": sheet_name,
                "source_row": r + 1,
                "extraction_method": "hydraulics_structured_table",
            }
        )
    return links

def extract_workbook(ctx: PipelineContext) -> Tuple[List[dict], List[dict], List[dict], List[dict]]:
    workbook_path = required_paths(ctx)["workbook"]
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    inventory_rows: List[dict] = []
    nodes_rows: List[dict] = []
    links_rows: List[dict] = []
    rational_rows: List[dict] = []

    keywords = ctx.extraction_cfg["classification_keywords"]
    scan_rows = int(ctx.extraction_cfg.get("candidate_header_scan_rows", 20))

    for sheet in wb.worksheets:
        all_rows = list(sheet.iter_rows(values_only=True))
        if not all_rows:
            continue

        max_row = len(all_rows)
        max_col = max(len(r) for r in all_rows)

        header_row_idx = 1
        best_count = -1
        best_headers: List[str] = []
        for i in range(1, min(max_row, scan_rows) + 1):
            vals = [normalize_header(v) for v in all_rows[i - 1]]
            count = sum(1 for v in vals if v)
            if count > best_count:
                best_count = count
                header_row_idx = i
                best_headers = vals

        inventory_rows.append(
            {
                "sheet_name": sheet.title,
                "rows": max_row,
                "cols": max_col,
                "detected_header_row": header_row_idx,
                "non_empty_headers": best_count,
                "headers_preview": " | ".join([h for h in best_headers if h][:12]),
            }
        )

        headers = [h if h else f"col_{idx+1}" for idx, h in enumerate(best_headers)]

        for offset, row in enumerate(all_rows[header_row_idx:], start=header_row_idx + 1):
            values = list(row)
            if not any(v is not None and str(v).strip() != "" for v in values):
                continue
            rec = row_to_dict(headers, values)
            rec["source_sheet"] = sheet.title
            rec["source_row"] = offset
            group = classify_record(rec, keywords)
            if group == "nodes":
                nodes_rows.append(rec)
            elif group == "links":
                links_rows.append(rec)
            elif group == "rational":
                rational_rows.append(rec)

        if sheet.title.strip().upper() == "HYDRAULICS":
            links_rows.extend(_extract_hydraulic_links_from_sheet(all_rows, sheet.title))

    return inventory_rows, nodes_rows, links_rows, rational_rows


def extract_pdf_metadata(ctx: PipelineContext) -> List[dict]:
    rows: List[dict] = []
    for p in required_paths(ctx)["pdfs"]:
        rows.append(
            {
                "path": str(p.relative_to(ROOT)),
                "exists": p.exists(),
                "size_bytes": p.stat().st_size if p.exists() else None,
                "sha256": sha256_file(p) if p.exists() else None,
                "note": "PDF parsing pass-through in current iteration",
            }
        )
    return rows


def run_extract(ctx: PipelineContext) -> StageResult:
    preflight = run_preflight(ctx)
    if preflight.status == "blocked":
        return StageResult("extract", "blocked", "Preflight blocked extraction due to missing inputs", preflight.produced_files)

    inventory, nodes, links, rational = extract_workbook(ctx)
    pdf_meta = extract_pdf_metadata(ctx)

    outputs = {
        "inventory": ROOT / "data/interim/workbook_sheet_inventory.csv",
        "nodes": ROOT / "data/interim/nodes_raw.csv",
        "links": ROOT / "data/interim/links_raw.csv",
        "rational": ROOT / "data/interim/rational_raw.csv",
        "pdf": ROOT / "data/interim/pdf_metadata.csv",
    }
    for k, path in outputs.items():
        if k == "inventory":
            write_csv(path, inventory)
        elif k == "nodes":
            write_csv(path, nodes)
        elif k == "links":
            write_csv(path, links)
        elif k == "rational":
            write_csv(path, rational)
        elif k == "pdf":
            write_csv(path, pdf_meta)

    summary = {
        "generated_at_utc": now_utc(),
        "counts": {
            "inventory_rows": len(inventory),
            "nodes_raw": len(nodes),
            "links_raw": len(links),
            "rational_raw": len(rational),
            "pdf_files": len(pdf_meta),
        },
    }
    summary_path = ROOT / "outputs/logs/extract_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    return StageResult(
        "extract",
        "ready",
        "Extraction completed (workbook discovered programmatically; PDFs metadata captured)",
        [str(p.relative_to(ROOT)) for p in list(outputs.values()) + [summary_path]],
    )


def canonicalize_rows(rows: List[dict]) -> List[dict]:
    canon: List[dict] = []
    for r in rows:
        nr = {}
        for k, v in r.items():
            nk = k.strip().lower().replace(" ", "_").replace("/", "_")
            nr[nk] = v
        canon.append(nr)
    return canon


def run_transform(ctx: PipelineContext) -> StageResult:
    import pandas as pd

    interim = ROOT / "data/interim"
    processed = ROOT / "data/processed"
    processed.mkdir(parents=True, exist_ok=True)

    sources = {
        "nodes": interim / "nodes_raw.csv",
        "links": interim / "links_raw.csv",
        "rational": interim / "rational_raw.csv",
    }
    missing = [k for k, p in sources.items() if not p.exists()]
    if missing:
        return StageResult("transform", "blocked", f"Missing interim files: {missing}", [])

    nodes_df = pd.read_csv(sources["nodes"]) if sources["nodes"].stat().st_size > 0 else pd.DataFrame()
    links_df = pd.read_csv(sources["links"]) if sources["links"].stat().st_size > 0 else pd.DataFrame()
    rational_df = pd.read_csv(sources["rational"]) if sources["rational"].stat().st_size > 0 else pd.DataFrame()

    nodes = canonicalize_rows(nodes_df.to_dict(orient="records"))
    links = canonicalize_rows(links_df.to_dict(orient="records"))
    rational = canonicalize_rows(rational_df.to_dict(orient="records"))

    tc_default = float(ctx.defaults_cfg["subcatchment"]["tc_length_default_ft"])
    min_width = float(ctx.defaults_cfg["subcatchment"]["min_width_ft"])
    imperv = float(ctx.defaults_cfg["subcatchment"]["percent_impervious_default"])
    slope = float(ctx.defaults_cfg["subcatchment"]["slope_percent_default"])

    subcatchment_defaults: List[dict] = []
    for rec in rational:
        area = _to_float(rec.get("area")) or _to_float(rec.get("area_ac"))
        q = _to_float(rec.get("q"))
        if area is None and q is not None and q > 0:
            area = max(0.05, min(5.0, q / 5.0))
        area = area or 0.0
        tc_len = _to_float(rec.get("tclength")) or _to_float(rec.get("tc_length")) or tc_default
        width = max((2.0 * area * 43560.0) / tc_len if tc_len > 0 else min_width, min_width)
        subcatchment_defaults.append(
            {
                "source_sheet": rec.get("source_sheet"),
                "source_row": rec.get("source_row"),
                "area_ac": round(area, 4),
                "tc_length_ft": round(tc_len, 3),
                "width_ft": round(width, 3),
                "slope_percent": slope,
                "percent_impervious": imperv,
            }
        )

    outputs = {
        "nodes": processed / "nodes.csv",
        "links": processed / "links.csv",
        "rational": processed / "rational_data.csv",
        "subs": processed / "subcatchment_defaults.csv",
    }
    write_csv(outputs["nodes"], nodes)
    write_csv(outputs["links"], links)
    write_csv(outputs["rational"], rational)
    write_csv(outputs["subs"], subcatchment_defaults)

    return StageResult("transform", "ready", "Transform completed", [str(p.relative_to(ROOT)) for p in outputs.values()])


def _build_runnable_model(ctx: PipelineContext) -> Tuple[str, dict]:
    import pandas as pd

    workbook_path = ROOT / "data/raw/Copy of BX-HH-Vista Ranch_10-30-2025.xlsm"
    if not workbook_path.exists():
        raise FileNotFoundError("Workbook missing")

    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    hyd = wb["HYDROLOGY"]
    di = wb["DI TABLE"]
    roughness = float(ctx.defaults_cfg["hydraulics"].get("manning_n_pipe_default", 0.013))
    min_slope = 0.002

    def norm_id(v: object) -> str | None:
        if v is None:
            return None
        t = str(v).strip()
        if not t or t.lower() in {"nan", "none"}:
            return None
        m = re.search(r"[-+]?\d+(?:\.\d+)?", t)
        if not m:
            return None
        x = float(m.group(0))
        return str(int(x)) if x.is_integer() else str(x).rstrip("0").rstrip(".")

    def as_float(v: object) -> float | None:
        try:
            f = float(str(v).strip())
            return f if math.isfinite(f) else None
        except Exception:
            return None

    def j_name(v: str) -> str:
        return f"J_{v.replace('.', '_')}"

    def in_name(v: str) -> str:
        return f"IN_{v.replace('.', '_')}"

    def out_name(v: str) -> str:
        return f"O_{v.replace('.', '_')}"

    assumption_apps: list[dict] = []
    def add_assumption(swmm_id: str, object_type: str, field_name: str, value: object, rule: str, context: str) -> None:
        assumption_apps.append({
            "swmm_object_id": swmm_id,
            "object_type": object_type,
            "field_name": field_name,
            "assumed_value": value,
            "assumption_rule": rule,
            "source_context": context,
        })

    # DI TABLE geometry
    di_geom: dict[str, dict] = {}
    for r_idx, r in enumerate(di.iter_rows(min_row=1, max_col=9, values_only=True), start=1):
        inlet_id = norm_id(r[1] if len(r) > 1 else None)
        if not inlet_id:
            continue
        di_geom[inlet_id] = {
            "up_inv": as_float(r[6] if len(r) > 6 else None),
            "dn_inv": as_float(r[7] if len(r) > 7 else None),
            "length": as_float(r[8] if len(r) > 8 else None),
            "source_row": r_idx,
        }

    # Parse HYDROLOGY + BASIN segmentation
    basins: list[dict] = []
    all_structures: list[dict] = []
    cur = {"index": 1, "junction_rows": [], "inlet_rows": []}
    for row_idx, row in enumerate(hyd.iter_rows(min_row=1, max_col=29, values_only=True), start=1):
        b = row[1] if len(row) > 1 else None  # B
        c = row[2] if len(row) > 2 else None  # C
        d = row[3] if len(row) > 3 else None  # D
        r = row[17] if len(row) > 17 else None  # R
        s_col = row[18] if len(row) > 18 else None  # S
        v_col = row[21] if len(row) > 21 else None  # V
        aa = row[26] if len(row) > 26 else None  # AA
        ac = row[28] if len(row) > 28 else None  # AC

        if str(b or "").strip().upper() == "BASIN":
            if cur["junction_rows"] or cur["inlet_rows"]:
                cur["basin_break_row"] = row_idx
                basins.append(cur)
            cur = {"index": cur["index"] + 1, "junction_rows": [], "inlet_rows": []}
            continue

        jct_id = norm_id(b)
        inlet_id = norm_id(c)
        rec = {
            "row_idx": row_idx,
            "jct_id": jct_id,
            "inlet_id": inlet_id,
            "known_q_cfs": as_float(r),
            "dia_in": as_float(s_col),
            "length_ft": as_float(v_col),
            "rim": as_float(aa),
            "soffit": as_float(ac),
            "invert": None,
            "col_d_value": d,
            "basin_index": cur["index"],
        }
        if rec["soffit"] is not None and rec["dia_in"] is not None:
            rec["invert"] = rec["soffit"] - (rec["dia_in"] / 12.0)

        is_j = bool(jct_id)
        is_i = bool(inlet_id) and not is_j
        if is_j or is_i:
            rec["structure_type"] = "junction" if is_j else "inlet"
            all_structures.append(rec)
        if is_j:

        is_junction_candidate = bool(jct_id)
        is_inlet_candidate = bool(inlet_id) and not bool(jct_id)
        rec["basin_index"] = cur["index"]
        if is_junction_candidate or is_inlet_candidate:
            rec["structure_type"] = "junction" if is_junction_candidate else "inlet"
            all_structures.append(rec)

        if is_junction_candidate:
            cur["junction_rows"].append(rec)
        elif is_i:
            cur["inlet_rows"].append(rec)
    if cur["junction_rows"] or cur["inlet_rows"]:
        basins.append(cur)

    def _structure_swmm_id(rec: dict) -> str:
        return j_name(rec["jct_id"]) if rec["structure_type"] == "junction" else in_name(rec["inlet_id"])

    def _completeness_score(rec: dict) -> int:
        if rec["structure_type"] == "junction":
            return sum(1 for k in ("rim", "soffit", "dia_in", "length_ft", "invert") if rec.get(k) is not None)
        return sum(1 for k in ("known_q_cfs",) if rec.get(k) is not None) + sum(
            1 for k in ("up_inv", "dn_inv", "length") if di_geom.get(rec["inlet_id"], {}).get(k) is not None
        )

    duplicate_groups: dict[tuple[int, str, str], list[dict]] = defaultdict(list)
    for basin in basins:
        for rec in basin["junction_rows"]:
            duplicate_groups[(basin["index"], "junction", _structure_swmm_id({**rec, "structure_type": "junction"}))].append(rec)
        for rec in basin["inlet_rows"]:
            duplicate_groups[(basin["index"], "inlet", _structure_swmm_id({**rec, "structure_type": "inlet"}))].append(rec)

    dropped_duplicates: list[dict] = []
    for basin in basins:
        for key_name, skey in (("junction_rows", "junction"), ("inlet_rows", "inlet")):
            resolved: list[dict] = []
            seen_ids: set[str] = set()
            for rec in basin[key_name]:
                sw = _structure_swmm_id({**rec, "structure_type": skey})
                if sw in seen_ids:
                    continue
                grp = [r for r in basin[key_name] if _structure_swmm_id({**r, "structure_type": skey}) == sw]
                if len(grp) == 1:
                    resolved.append(rec)
                    seen_ids.add(sw)
                    continue
                grp_sorted = sorted(grp, key=lambda r: (-_completeness_score({**r, "structure_type": skey}), r["row_idx"]))
                keep = grp_sorted[0]
                resolved.append(keep)
                seen_ids.add(sw)
                for drop in grp_sorted[1:]:
                    dropped_duplicates.append({
                        "basin_index": basin["index"],
                        "structure_type": skey,
                        "swmm_id": sw,
                        "kept_source_row": keep["row_idx"],
                        "dropped_source_row": drop["row_idx"],
                        "reason": "duplicate_id_lower_completeness_or_later_row",
                    })
            basin[key_name] = sorted(resolved, key=lambda r: r["row_idx"])

    # duplicate/conflict map (post-dedup awareness)
    id_counts = defaultdict(int)
    for rec in all_structures:
        rid = rec["jct_id"] if rec["structure_type"] == "junction" else rec["inlet_id"]
        id_counts[(rec["structure_type"], str(rid))] += 1
    # duplicate/conflict map
    id_counts = defaultdict(int)
    for rec in all_structures:
        id_counts[(rec["structure_type"], rec["source_id"] if "source_id" in rec else (rec["jct_id"] if rec["structure_type"]=="junction" else rec["inlet_id"]))] += 1

    def interpolate_series(vals: list[float | None], lens: list[float | None]) -> list[float | None]:
        out = vals[:]
        n = len(out)
        known = [i for i,v in enumerate(out) if v is not None]
        if not known:
            return out
        for i in range(n):
            if out[i] is not None:
                continue
            up = max([k for k in known if k < i], default=None)
            dn = min([k for k in known if k > i], default=None)
            if up is not None and dn is not None:
                seg = [max((lens[k] or 0.0), 1.0) for k in range(up, dn)]
                tot = sum(seg)
                if tot <= 0:
                    frac = (i-up)/(dn-up)
                else:
                    frac = sum(seg[: i-up]) / tot
                out[i] = out[up] + (out[dn] - out[up]) * frac
            elif up is not None:
                dist = sum(max((lens[k] or 0.0), 1.0) for k in range(up, i))
                out[i] = out[up] - min_slope * dist
            elif dn is not None:
                dist = sum(max((lens[k] or 0.0), 1.0) for k in range(i, dn))
                out[i] = out[dn] + min_slope * dist
        return out

    def build_network(apply_assumptions: bool):
        nodes: dict[str, dict] = {}
        outfalls: dict[str, dict] = {}
        conduits: list[dict] = []
        inflow_by_node: defaultdict[str, float] = defaultdict(float)
        findings: list[dict] = []
        crosswalk: list[dict] = []
        seen_cids: set[str] = set()

        def add_finding(sev: str, check: str, msg: str, source_row: int | None, entity: str) -> None:
            findings.append({"severity": sev, "check_name": check, "message": msg, "source_row": source_row, "entity": entity})

        for basin in basins:
            jrows = basin["junction_rows"]
            if not jrows:
                continue

            # assumption interpolation for junction inverts/rims
            if apply_assumptions:
                inverts = [jr.get("invert") for jr in jrows]
                rims = [jr.get("rim") for jr in jrows]
                lens = [jr.get("length_ft") for jr in jrows]
                inf = interpolate_series(inverts, lens)
                rim_interp = interpolate_series(rims, lens)
                for i, jr in enumerate(jrows):
                    sid = j_name(jr["jct_id"])
                    if jr.get("invert") is None and inf[i] is not None:
                        jr["invert"] = inf[i]
                        add_assumption(sid, "node", "invert", round(inf[i], 6), "junction_linear_interpolation_or_min_slope", f"HYDROLOGY row {jr['row_idx']}")
                    if jr.get("rim") is None:
                        if rim_interp[i] is not None:
                            jr["rim"] = rim_interp[i]
                            add_assumption(sid, "node", "rim", round(rim_interp[i], 6), "junction_rim_interpolation", f"HYDROLOGY row {jr['row_idx']}")
                        elif jr.get("invert") is not None:
                            jr["rim"] = jr["invert"] + 8.0
                            add_assumption(sid, "node", "rim", round(jr["rim"], 6), "junction_rim_default_plus_8ft", f"HYDROLOGY row {jr['row_idx']}")
        # inlet rows are attached to most-recent upstream junction in same basin
        pending_inlets = list(basin["inlet_rows"])
        for inlet in pending_inlets:
            recv = next((jr for jr in reversed(jrows) if jr["row_idx"] < inlet["row_idx"]), None)
            if recv is None:
                add_finding("HIGH", "inlet_receiving_junction_missing", "No upstream receiving junction found for inlet row.", inlet["row_idx"], in_name(inlet["inlet_id"]))
                continue
            inlet_swmm = in_name(inlet["inlet_id"])
            recv_swmm = j_name(recv["jct_id"])
            recv_node = nodes.get(recv_swmm)
            if recv_node is None:
                add_finding("HIGH", "inlet_receiving_junction_missing", "Receiving junction missing for inlet row.", inlet["row_idx"], inlet_swmm)
                continue

            for jr in jrows:
                jid = j_name(jr["jct_id"])
                if jr.get("invert") is None or jr.get("rim") is None:
                    add_finding("HIGH", "junction_missing_elevation", "Missing junction invert/rim from HYDROLOGY AC/AA.", jr["row_idx"], jid)
                    continue
                nodes[jid] = {
                    "elev": jr["invert"],
                    "max_depth": max(jr["rim"] - jr["invert"], 0.01),
                    "source_type": "junction",
                    "source_row": jr["row_idx"],
                    "source_id": jr["jct_id"],
                    "rim": jr["rim"],
                    "basin_index": basin["index"],
                }
                crosswalk.append({"object_type": "node", "swmm_id": jid, "source_tab": "HYDROLOGY", "source_row": jr["row_idx"], "source_id": jr["jct_id"], "notes": "junction from HYDROLOGY B"})

            # inlets
            for inlet in basin["inlet_rows"]:
                inlet_swmm = in_name(inlet["inlet_id"])
                recv = next((jr for jr in reversed(jrows) if jr["row_idx"] < inlet["row_idx"]), None)
                recv_swmm = j_name(recv["jct_id"]) if recv else None
                recv_node = nodes.get(recv_swmm) if recv_swmm else None
                if recv_node is None:
                    add_finding("HIGH", "inlet_receiving_junction_missing", "No upstream receiving junction found for inlet row.", inlet["row_idx"], inlet_swmm)
                    continue
                di_rec = di_geom.get(inlet["inlet_id"], {})
                up_inv = di_rec.get("up_inv")
                dn_inv = di_rec.get("dn_inv")
                length_ft = di_rec.get("length")
                if apply_assumptions:
                    if length_ft is None:
                        length_ft = 25.0
                        add_assumption(inlet_swmm, "link", "length_ft", length_ft, "inlet_length_default_25ft", f"HYDROLOGY row {inlet['row_idx']};DI missing I")
                    if dn_inv is None and recv_node.get("elev") is not None:
                        dn_inv = recv_node["elev"]
                        add_assumption(inlet_swmm, "link", "dn_invert", round(dn_inv,6), "inlet_dn_invert_from_receiving_junction", f"HYDROLOGY row {inlet['row_idx']}")
                    if up_inv is None and dn_inv is not None:
                        up_inv = dn_inv + min_slope * length_ft
                        add_assumption(inlet_swmm, "node", "invert", round(up_inv,6), "inlet_up_invert_from_min_slope", f"HYDROLOGY row {inlet['row_idx']}")
                    if dn_inv is None and up_inv is not None:
                        dn_inv = up_inv - min_slope * length_ft
                        add_assumption(inlet_swmm, "link", "dn_invert", round(dn_inv,6), "inlet_dn_invert_from_up_min_slope", f"HYDROLOGY row {inlet['row_idx']}")
                if up_inv is None or dn_inv is None or length_ft is None:
                    add_finding("HIGH", "inlet_di_geometry_missing", "Missing DI TABLE geometry for inlet->junction conduit.", inlet["row_idx"], inlet_swmm)
                    continue

                if inlet_swmm not in nodes:
                    rim_assumed = recv_node["rim"]
                    nodes[inlet_swmm] = {
                        "elev": up_inv,
                        "max_depth": max(rim_assumed - up_inv, 0.01),
                        "source_type": "inlet",
                        "source_row": inlet["row_idx"],
                        "source_id": inlet["inlet_id"],
                        "rim": rim_assumed,
                        "basin_index": basin["index"],
                    }
                    crosswalk.append({"object_type": "node", "swmm_id": inlet_swmm, "source_tab": "HYDROLOGY/DI TABLE", "source_row": inlet["row_idx"], "source_id": inlet["inlet_id"], "notes": "inlet node; rim assumed equal to receiving junction rim"})
                    if di_geom.get(inlet["inlet_id"], {}).get("up_inv") is None and apply_assumptions:
                        add_assumption(inlet_swmm, "node", "rim", rim_assumed, "inlet_rim_equals_receiving_junction", f"HYDROLOGY row {inlet['row_idx']}")

                cid = f"L_{inlet_swmm}__{recv_swmm}"
                if cid not in seen_cids:
                    seen_cids.add(cid)
                    conduits.append({"cid": cid, "us": inlet_swmm, "ds": recv_swmm, "length": max(length_ft, 0.1), "dia_ft": 1.0, "source_row": inlet["row_idx"], "source_tab": "DI TABLE", "link_type": "inlet", "basin_index": basin["index"]})
                    crosswalk.append({"object_type": "conduit", "swmm_id": cid, "source_tab": "DI TABLE", "source_row": inlet["row_idx"], "source_id": inlet["inlet_id"], "notes": "inlet->junction conduit"})
                    if apply_assumptions and di_rec.get("length") is None:
                        add_assumption(cid, "link", "diameter_in", 12, "inlet_conduit_diameter_default_12in", f"HYDROLOGY row {inlet['row_idx']}")

                if inlet.get("known_q_cfs") is not None and inlet["known_q_cfs"] > 0:
                    inflow_by_node[inlet_swmm] += inlet["known_q_cfs"]
                    crosswalk.append({"object_type": "inflow", "swmm_id": inlet_swmm, "source_tab": "HYDROLOGY", "source_row": inlet["row_idx"], "source_id": inlet["inlet_id"], "notes": f"known flow cfs={inlet['known_q_cfs']}"})
                elif inlet.get("known_q_cfs") is None:
                    add_finding("HIGH", "inlet_known_flow_missing", "Inlet known flow missing/invalid in HYDROLOGY col R.", inlet["row_idx"], inlet_swmm)

            # j->j conduits
            for i, jr in enumerate(jrows[:-1]):
                dn = jrows[i + 1]
                us_swmm = j_name(jr["jct_id"])
                dn_swmm = j_name(dn["jct_id"])
                if us_swmm == dn_swmm:
                    add_finding("HIGH", "junction_self_loop_candidate", "Self-loop prevented during build.", jr["row_idx"], us_swmm)
                    continue
                if us_swmm not in nodes or dn_swmm not in nodes:
                    add_finding("HIGH", "junction_conduit_endpoint_missing_node", "Skipped conduit because endpoint node missing due incomplete junction geometry.", jr["row_idx"], f"{us_swmm}->{dn_swmm}")
                    continue
                if jr.get("length_ft") is None or jr.get("dia_in") is None:
                    add_finding("HIGH", "junction_pipe_geometry_missing", "Missing S/V geometry for junction downstream conduit.", jr["row_idx"], us_swmm)
                    continue
                cid = f"L_{us_swmm}__{dn_swmm}"
                if cid in seen_cids:
                    continue
                seen_cids.add(cid)
                conduits.append({"cid": cid, "us": us_swmm, "ds": dn_swmm, "length": max(jr["length_ft"], 0.1), "dia_ft": max(jr["dia_in"] / 12.0, 0.5), "source_row": jr["row_idx"], "source_tab": "HYDROLOGY", "link_type": "junction", "basin_index": basin["index"]})
                crosswalk.append({"object_type": "conduit", "swmm_id": cid, "source_tab": "HYDROLOGY", "source_row": jr["row_idx"], "source_id": jr["jct_id"], "notes": "junction->junction downstream conduit"})

            # terminal to outfall
            terminal = jrows[-1]
            term_swmm = j_name(terminal["jct_id"])
            of_swmm = out_name(terminal["jct_id"])
            if term_swmm not in nodes and apply_assumptions:
                up = next((j for j in reversed(jrows[:-1]) if j_name(j["jct_id"]) in nodes), None)
                if up:
                    up_swmm = j_name(up["jct_id"])
                    l_up = max((up.get("length_ft") or 10.0), 10.0)
                    inv = nodes[up_swmm]["elev"] - min_slope * l_up
                    rim = inv + 8.0
                    nodes[term_swmm] = {"elev": inv, "max_depth": 8.0, "source_type": "junction", "source_row": terminal["row_idx"], "source_id": terminal["jct_id"], "rim": rim, "basin_index": basin["index"]}
                    add_assumption(term_swmm, "node", "invert", round(inv,6), "terminal_invert_from_upstream_min_slope", f"HYDROLOGY row {terminal['row_idx']}")
                    add_assumption(term_swmm, "node", "rim", round(rim,6), "terminal_rim_default_plus_8ft", f"HYDROLOGY row {terminal['row_idx']}")
                    crosswalk.append({"object_type": "node", "swmm_id": term_swmm, "source_tab": "HYDROLOGY", "source_row": terminal["row_idx"], "source_id": terminal["jct_id"], "notes": "junction assumed from terminal interpolation"})
                else:
                    fallback_inv = 0.0
                    if terminal.get("rim") is not None:
                        fallback_inv = terminal["rim"] - 8.0
                    nodes[term_swmm] = {"elev": fallback_inv, "max_depth": 8.0, "source_type": "junction", "source_row": terminal["row_idx"], "source_id": terminal["jct_id"], "rim": fallback_inv + 8.0, "basin_index": basin["index"]}
                    add_assumption(term_swmm, "node", "invert", round(fallback_inv, 6), "terminal_invert_default_no_upstream", f"HYDROLOGY row {terminal['row_idx']}")
                    add_assumption(term_swmm, "node", "rim", round(fallback_inv + 8.0, 6), "terminal_rim_default_plus_8ft", f"HYDROLOGY row {terminal['row_idx']}")
                    crosswalk.append({"object_type": "node", "swmm_id": term_swmm, "source_tab": "HYDROLOGY", "source_row": terminal["row_idx"], "source_id": terminal["jct_id"], "notes": "junction assumed from default terminal fallback"})

            if term_swmm not in nodes:
                add_finding("HIGH", "terminal_junction_missing_node", "Skipped basin outfall link because terminal junction node missing geometry.", terminal["row_idx"], term_swmm)
                continue

            l_out = max(10.0, terminal.get("length_ft") or 10.0)
            dia_in = terminal.get("dia_in")
            if dia_in is None and apply_assumptions:
                upstream_dia = None
                for c in reversed(conduits):
                    if c["ds"] == term_swmm and c.get("dia_ft"):
                        upstream_dia = c["dia_ft"] * 12.0
                        break
                if upstream_dia is not None:
                    dia_in = upstream_dia
                    add_assumption(f"L_{term_swmm}__{of_swmm}", "link", "diameter_in", round(dia_in,6), "terminal_diameter_from_upstream_conduit", f"HYDROLOGY row {terminal['row_idx']}")
                else:
                    dia_in = 18.0
                    add_assumption(f"L_{term_swmm}__{of_swmm}", "link", "diameter_in", dia_in, "terminal_diameter_default_18in", f"HYDROLOGY row {terminal['row_idx']}")

            if terminal.get("length_ft") is None and apply_assumptions:
                add_assumption(f"L_{term_swmm}__{of_swmm}", "link", "length_ft", l_out, "terminal_outfall_length_default_or_min10", f"HYDROLOGY row {terminal['row_idx']}")

            out_elev = nodes[term_swmm]["elev"] - min_slope * l_out
            outfalls[of_swmm] = {"elev": out_elev, "source_row": terminal["row_idx"], "terminal": term_swmm, "basin_index": basin["index"]}
            if apply_assumptions:
                add_assumption(of_swmm, "outfall", "invert", round(out_elev,6), "outfall_invert_from_terminal_min_slope", f"HYDROLOGY row {terminal['row_idx']}")
            cid = f"L_{term_swmm}__{of_swmm}"
            if cid not in seen_cids:
                seen_cids.add(cid)
                conduits.append({"cid": cid, "us": term_swmm, "ds": of_swmm, "length": l_out, "dia_ft": max((dia_in or 18.0)/12.0, 0.5), "source_row": terminal["row_idx"], "source_tab": "HYDROLOGY", "link_type": "outfall", "basin_index": basin["index"]})
                crosswalk.append({"object_type": "conduit", "swmm_id": cid, "source_tab": "HYDROLOGY", "source_row": terminal["row_idx"], "source_id": terminal["jct_id"], "notes": "terminal junction->outfall per BASIN rule"})

        conduits = [c for c in conduits if c["us"] != c["ds"]]
        return nodes, outfalls, conduits, inflow_by_node, findings, crosswalk

    # Phase A diagnostics run (no assumptions)
    base_nodes, base_outfalls, base_conduits, _base_inflows, base_findings, _base_crosswalk = build_network(False)

    base_node_ids = set(base_nodes.keys())
    missing_rows = []
    for basin in basins:
        jrows = basin.get("junction_rows", [])
        for idx, jr in enumerate(jrows):
            sw = j_name(jr["jct_id"])
            if sw in base_node_ids:
                conduits.append({
                    "cid": cid,
                    "us": inlet_swmm,
                    "ds": recv_swmm,
                    "length": max(length_ft, 0.1),
                    "dia_ft": max((12.0 / 12.0), 0.5),
                    "source_row": inlet["row_idx"],
                    "source_tab": "DI TABLE",
                    "link_type": "inlet",
                })
                crosswalk.append({"object_type": "conduit", "swmm_id": cid, "source_tab": "DI TABLE", "source_row": inlet["row_idx"], "source_id": inlet["inlet_id"], "notes": "inlet->junction conduit"})

            if inlet.get("known_q_cfs") is not None and inlet["known_q_cfs"] > 0:
                inflow_by_node[inlet_swmm] += inlet["known_q_cfs"]
                crosswalk.append({"object_type": "inflow", "swmm_id": inlet_swmm, "source_tab": "HYDROLOGY", "source_row": inlet["row_idx"], "source_id": inlet["inlet_id"], "notes": f"known flow cfs={inlet['known_q_cfs']}"})
            elif inlet.get("known_q_cfs") is None:
                add_finding("HIGH", "inlet_known_flow_missing", "Inlet known flow missing/invalid in HYDROLOGY col R.", inlet["row_idx"], inlet_swmm)

        # junction->downstream junction conduits from current row attributes
        for i, jr in enumerate(jrows[:-1]):
            dn = jrows[i + 1]
            us_swmm = j_name(jr["jct_id"])
            dn_swmm = j_name(dn["jct_id"])
            if us_swmm == dn_swmm:
                add_finding("HIGH", "junction_self_loop_candidate", "Self-loop prevented during build.", jr["row_idx"], us_swmm)
                continue
            if us_swmm not in nodes or dn_swmm not in nodes:
                add_finding("HIGH", "junction_conduit_endpoint_missing_node", "Skipped conduit because endpoint node missing due incomplete junction geometry.", jr["row_idx"], f"{us_swmm}->{dn_swmm}")
                continue
            if jr["length_ft"] is None or jr["dia_in"] is None:
                add_finding("HIGH", "junction_pipe_geometry_missing", "Missing S/V geometry for junction downstream conduit.", jr["row_idx"], us_swmm)
                continue
            cid = f"L_{us_swmm}__{dn_swmm}"
            if cid in seen_cids:
                continue
            req_missing = []
            req_present = []
            for fld, key in [("AA", "rim"), ("AC", "soffit"), ("S", "dia_in")]:
                if jr.get(key) is None:
                    req_missing.append(fld)
                else:
                    req_present.append(fld)
            dn_sw = j_name(jrows[idx + 1]["jct_id"]) if idx < len(jrows) - 1 else ""
            pattern = "missing " + "/".join(req_missing) if req_missing else "none"
            decision = "MISSING_DATA" if req_missing else "TOPOLOGY_AMBIGUITY"
            rec_fix = "interpolate invert/rim" if any(f in req_missing for f in ["AA", "AC", "S"]) else "review topology order"
            missing_rows.append({
                "source_tab": "HYDROLOGY",
                "source_row": jr["row_idx"],
                "basin_index": basin["index"],
                "structure_type": "junction",
                "source_id": jr["jct_id"],
                "swmm_id": sw,
                "receiving_junction_swmm_id": "",
                "downstream_junction_swmm_id": dn_sw,
                "required_fields_missing": ";".join(req_missing),
                "required_fields_present": ";".join(req_present),
                "decision_class": decision,
                "recommended_fix": rec_fix,
            })
        for inlet in basin.get("inlet_rows", []):
            sw = in_name(inlet["inlet_id"])
            if sw in base_node_ids:
                continue
            recv = next((jr for jr in reversed(jrows) if jr["row_idx"] < inlet["row_idx"]), None)
            recv_sw = j_name(recv["jct_id"]) if recv else ""
            di_rec = di_geom.get(inlet["inlet_id"], {})
            req_missing=[]; req_present=[]
            for fld,key in [("DI_G","up_inv"),("DI_H","dn_inv"),("DI_I","length")]:
                if di_rec.get(key) is None: req_missing.append(fld)
                else: req_present.append(fld)
            decision = "TOPOLOGY_AMBIGUITY" if not recv_sw else ("MISSING_DATA" if req_missing else "MISSING_DATA")
            rec_fix = "resolve receiving junction" if not recv_sw else "DI TABLE missing length/invert; apply min-slope"
            missing_rows.append({
                "source_tab":"HYDROLOGY",
                "source_row":inlet["row_idx"],
                "basin_index":basin["index"],
                "structure_type":"inlet",
                "source_id":inlet["inlet_id"],
                "swmm_id":sw,
                "receiving_junction_swmm_id":recv_sw,
                "downstream_junction_swmm_id":"",
                "required_fields_missing":";".join(req_missing),
                "required_fields_present":";".join(req_present),
                "decision_class":decision,
                "recommended_fix":rec_fix,
            })

    # flag duplicates/conflicts
    for r in missing_rows:
        k = (r["structure_type"], str(r["source_id"]))
        if id_counts.get(k, 0) > 1:
            r["decision_class"] = "DUPLICATE_OR_CONFLICT"
            r["recommended_fix"] = "deduplicate conflicting IDs/rows"

    for d in dropped_duplicates:
        missing_rows.append({
            "source_tab": "HYDROLOGY",
            "source_row": d["dropped_source_row"],
            "basin_index": d["basin_index"],
            "structure_type": d["structure_type"],
            "source_id": d["swmm_id"].replace("J_", "").replace("IN_", "").replace("_", "."),
            "swmm_id": d["swmm_id"],
            "receiving_junction_swmm_id": "",
            "downstream_junction_swmm_id": "",
            "required_fields_missing": "",
            "required_fields_present": "",
            "decision_class": "DUPLICATE_OR_CONFLICT",
            "recommended_fix": f"drop duplicate row {d['dropped_source_row']}; keep row {d['kept_source_row']}",
        })

            })
        for inlet in basin.get("inlet_rows", []):
            sw = in_name(inlet["inlet_id"])
            if sw in base_node_ids:
                continue
            recv = next((jr for jr in reversed(jrows) if jr["row_idx"] < inlet["row_idx"]), None)
            recv_sw = j_name(recv["jct_id"]) if recv else ""
            di_rec = di_geom.get(inlet["inlet_id"], {})
            req_missing=[]; req_present=[]
            for fld,key in [("DI_G","up_inv"),("DI_H","dn_inv"),("DI_I","length")]:
                if di_rec.get(key) is None: req_missing.append(fld)
                else: req_present.append(fld)
            decision = "MISSING_DATA" if req_missing else "TOPOLOGY_AMBIGUITY"
            rec_fix = "DI TABLE missing length/invert; apply min-slope" if req_missing else "resolve receiving junction"
            missing_rows.append({
                "source_tab":"HYDROLOGY",
                "source_row":inlet["row_idx"],
                "basin_index":basin["index"],
                "structure_type":"inlet",
                "source_id":inlet["inlet_id"],
                "swmm_id":sw,
                "receiving_junction_swmm_id":recv_sw,
                "downstream_junction_swmm_id":"",
                "required_fields_missing":";".join(req_missing),
                "required_fields_present":";".join(req_present),
                "decision_class":decision,
                "recommended_fix":rec_fix,
                "link_type": "junction",
            })
            crosswalk.append({"object_type": "conduit", "swmm_id": cid, "source_tab": "HYDROLOGY", "source_row": jr["row_idx"], "source_id": jr["jct_id"], "notes": "junction->junction downstream conduit"})

        # terminal junction to basin-specific outfall
        terminal = jrows[-1]
        term_swmm = j_name(terminal["jct_id"])
        of_swmm = out_name(terminal["jct_id"])
        if term_swmm not in nodes:
            add_finding("HIGH", "terminal_junction_missing_node", "Skipped basin outfall link because terminal junction node missing geometry.", terminal["row_idx"], term_swmm)
            continue
        outfalls[of_swmm] = {"elev": nodes[term_swmm]["elev"], "source_row": terminal["row_idx"], "terminal": term_swmm, "basin_index": basin["index"]}
        cid = f"L_{term_swmm}__{of_swmm}"
        if terminal.get("length_ft") is None or terminal.get("dia_in") is None:
            add_finding("HIGH", "terminal_outfall_geometry_missing", "Skipped terminal->outfall conduit due missing S/V geometry on terminal row.", terminal["row_idx"], cid)
            continue
        if cid not in seen_cids:
            seen_cids.add(cid)
            conduits.append({
                "cid": cid,
                "us": term_swmm,
                "ds": of_swmm,
                "length": max(terminal["length_ft"], 0.1),
                "dia_ft": max(terminal["dia_in"] / 12.0, 0.5),
                "source_row": terminal["row_idx"],
                "source_tab": "HYDROLOGY",
                "link_type": "outfall",
            })

    # flag duplicates/conflicts
    for r in missing_rows:
        k = (r["structure_type"], str(r["source_id"]))
        if id_counts.get(k, 0) > 1:
            r["decision_class"] = "DUPLICATE_OR_CONFLICT"
            r["recommended_fix"] = "deduplicate conflicting IDs/rows"

    md_df = pd.DataFrame(missing_rows)
    md_df.to_csv(ROOT / "outputs/review/missing_data_diagnostics.csv", index=False)

    # summary markdown
    lines = ["# Missing Data Summary", ""]
    if not md_df.empty:
        lines.append("## Counts by basin + structure type + decision class")
        grp = md_df.groupby(["basin_index", "structure_type", "decision_class"]).size().reset_index(name="count")
        for _, r in grp.iterrows():
            lines.append(f"- Basin {int(r['basin_index'])} | {r['structure_type']} | {r['decision_class']}: {int(r['count'])}")
        lines.append("## Counts by basin and structure type")
        grp = md_df.groupby(["basin_index", "structure_type"]).size().reset_index(name="count")
        for _, r in grp.iterrows():
            lines.append(f"- Basin {int(r['basin_index'])} | {r['structure_type']}: {int(r['count'])}")
        lines.append("")
        lines.append("## Top 10 missing-field patterns")
        pat = md_df["required_fields_missing"].fillna("").replace("", "none").value_counts().head(10)
        for p, c in pat.items():
            lines.append(f"- {p}: {int(c)}")
        lines.append("")
        lines.append("## 'none' missing-field pattern by decision class")
        none_df = md_df[md_df["required_fields_missing"].fillna("") == ""]
        if none_df.empty:
            lines.append("- none: 0")
        else:
            for cls, c in none_df["decision_class"].value_counts().items():
                lines.append(f"- none | {cls}: {int(c)}")
        lines.append("")
        lines.append("## Terminal outfall failures (pre-assumption)")
        for basin in basins:
            if not basin.get("junction_rows"):
                continue
            term = basin["junction_rows"][-1]
            tid = j_name(term["jct_id"])
            oid = out_name(term["jct_id"])
            ok = oid in base_outfalls and any(c["ds"] == oid for c in base_conduits)
            if not ok:
                miss = []
                if term.get("rim") is None: miss.append("AA")
                if term.get("soffit") is None: miss.append("AC")
                if term.get("dia_in") is None: miss.append("S")
                if term.get("length_ft") is None: miss.append("V")
                lines.append(f"- Basin {basin['index']} terminal {tid} -> {oid}: missing {', '.join(miss) if miss else 'topology/other'}")
    (ROOT / "outputs/review/missing_data_summary.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    assumptions_enabled = os.getenv("PIPELINE_ENABLE_ASSUMPTIONS", "0") == "1"
    nodes, outfalls, conduits, inflow_by_node, findings, crosswalk = build_network(assumptions_enabled)

    # coordinates: orthogonal schematic from lengths
    # deterministic orthogonal schematic coordinates from link lengths
    coords: dict[str, tuple[float, float]] = {}
    x0 = 0.0
    y0 = 0.0
    for basin in basins:
        if not basin.get("junction_rows"):
            continue
        term = basin["junction_rows"][-1]
        t_sw = j_name(term["jct_id"])
        if t_sw not in nodes:
            continue
        o_sw = out_name(term["jct_id"])
        coords[t_sw] = (x0, y0)
        coords[o_sw] = (x0, y0 - 10.0)
        jset = {j_name(jr["jct_id"]) for jr in basin["junction_rows"] if j_name(jr["jct_id"]) in nodes}
        incoming_j = defaultdict(list)
        incoming_i = defaultdict(list)
        for c in conduits:
            if c["ds"] in jset:
                if c.get("link_type") == "junction":
                    incoming_j[c["ds"]].append(c)
                if c.get("link_type") == "inlet":
                    incoming_i[c["ds"]].append(c)
        occupied = {coords[t_sw]: t_sw, coords[o_sw]: o_sw}

        def place(node_id: str, recv_id: str, axis: str, sign: int, L: float):
            rx, ry = coords[recv_id]
            px = rx + (sign * L if axis == "x" else 0.0)
            py = ry + (sign * L if axis == "y" else 0.0)
            k = 0
            while (px, py) in occupied and occupied[(px, py)] != node_id:
                k += 1
                if axis == "x": py = ry + 10.0 * k
                else: px = rx + 10.0 * k
                findings.append({"severity": "LOW", "check_name": "coordinate_collision_resolved", "message": "Coordinate collision resolved", "source_row": None, "entity": node_id})
            coords[node_id] = (px, py)
            occupied[(px, py)] = node_id
        terminal = basin["junction_rows"][-1]
        term_swmm = j_name(terminal["jct_id"])
        if term_swmm not in nodes:
            continue
        coords[term_swmm] = (x0, y0)
        of_swmm = out_name(terminal["jct_id"])
        coords[of_swmm] = (x0, y0 - 10.0)

        jset = {j_name(jr["jct_id"]) for jr in basin["junction_rows"]}
        incoming_j = defaultdict(list)
        incoming_i = defaultdict(list)
        for c in conduits:
            if c["ds"] not in jset:
                continue
            if c.get("link_type") == "junction":
                incoming_j[c["ds"]].append(c)
            elif c.get("link_type") == "inlet":
                incoming_i[c["ds"]].append(c)

        occupied: dict[tuple[float, float], str] = {coords[term_swmm]: term_swmm, coords[of_swmm]: of_swmm}

        def place_node(node_id: str, recv_id: str, axis: str, sign: int, length: float) -> None:
            rx, ry = coords[recv_id]
            px = rx + (sign * length if axis == "x" else 0.0)
            py = ry + (sign * length if axis == "y" else 0.0)
            key = (px, py)
            k = 0
            while key in occupied and occupied[key] != node_id:
                k += 1
                if axis == "x":
                    key = (px, py + 10.0 * k)
                else:
                    key = (px + 10.0 * k, py)
                add_finding("LOW", "coordinate_collision_resolved", f"Coordinate collision resolved against {occupied.get((px, py), 'unknown')}", None, node_id)
            coords[node_id] = key
            occupied[key] = node_id

        changed = True
        while changed:
            changed = False
            for recv in list(jset):
                if recv not in coords:
                    continue
                for c in incoming_j.get(recv, []):
                    if c["us"] in nodes and c["us"] not in coords:
                        place(c["us"], recv, "y", +1, c["length"])
                        changed = True
                ins = [c for c in incoming_i.get(recv, []) if c["us"] in nodes and c["us"] not in coords]
                has_j = bool(incoming_j.get(recv))
                if not ins:
                    continue
                if has_j:
                    slot = [("x", -1), ("x", +1), ("y", -1)]
                elif len(ins) == 1:
                    slot = [("y", +1)]
                elif len(ins) == 2:
                    slot = [("x", -1), ("x", +1)]
                else:
                    slot = [("x", +1), ("y", +1), ("x", -1), ("y", -1)]
                for i, c in enumerate(ins):
                    ax, sg = slot[i % len(slot)]
                    place(c["us"], recv, ax, sg, c["length"])
                    changed = True

        xs = [xy[0] for nid, xy in coords.items() if nid in jset or nid.startswith("IN_") or nid == o_sw]
        max_x = max(xs) if xs else x0
        max_len = max([c["length"] for c in conduits if c.get("basin_index") == basin["index"]] or [100.0])
        x0 = max_x + max(200.0, 2 * max_len)

    # outputs
                if recv in incoming_j:
                    for c in incoming_j[recv]:
                        if c["us"] in nodes and c["us"] not in coords:
                            place_node(c["us"], recv, "y", +1, c["length"])
                            changed = True
                ins = [c for c in incoming_i.get(recv, []) if c["us"] in nodes and c["us"] not in coords]
                if not ins:
                    continue
                has_j = any(c["us"] in coords for c in incoming_j.get(recv, [])) or bool(incoming_j.get(recv))
                if has_j:
                    side_cycle = [("x", +1), ("x", -1), ("y", -1)]
                elif len(ins) == 1:
                    side_cycle = [("y", +1)]
                elif len(ins) == 2:
                    side_cycle = [("x", -1), ("x", +1)]
                else:
                    side_cycle = [("x", +1), ("y", +1), ("x", -1), ("y", -1)]
                for idx, c in enumerate(ins):
                    axis, sign = side_cycle[idx % len(side_cycle)]
                    place_node(c["us"], recv, axis, sign, c["length"])
                    changed = True

        xs = [p[0] for n, p in coords.items() if n in jset or n.startswith("IN_") or n == of_swmm]
        max_x = max(xs) if xs else x0
        max_len = max([c["length"] for c in conduits if c.get("link_type") in {"junction", "inlet", "outfall"}] or [100.0])
        x0 = max_x + max(200.0, 2 * max_len)

    # write traceability artifacts
    pd.DataFrame(crosswalk).to_csv(ROOT / "outputs/review/source_traceability_crosswalk.csv", index=False)
    pd.DataFrame(findings).to_csv(ROOT / "outputs/qa/inlet_knownflow_hydraulic_baseline_findings.csv", index=False)
    pd.DataFrame(assumption_apps).to_csv(ROOT / "outputs/review/assumption_applications.csv", index=False)

    modeled_nodes = set(nodes.keys())
    assumption_swmm_ids = {str(a.get("swmm_object_id")) for a in assumption_apps}
    dropped_swmm_ids = {d["swmm_id"] for d in dropped_duplicates}
    by_swmm: dict[str, dict] = {}
    for rec in all_structures:
        sid = rec["jct_id"] if rec["structure_type"] == "junction" else rec["inlet_id"]
        swmm_id = j_name(sid) if rec["structure_type"] == "junction" else in_name(sid)
        prev = by_swmm.get(swmm_id)
        if prev is None or rec["row_idx"] < prev["source_row"]:
            by_swmm[swmm_id] = {
                "source_tab": "HYDROLOGY",
                "source_row": rec["row_idx"],
                "basin_index": rec["basin_index"],
                "structure_type": rec["structure_type"],
                "source_id": sid,
                "swmm_id": swmm_id,
            }

    unaccounted_rows = []
    for swmm_id, row in sorted(by_swmm.items(), key=lambda kv: kv[0]):
        modeled = swmm_id in modeled_nodes
        if modeled:
            modeled_reason = "MODELED_WITH_ASSUMPTION" if swmm_id in assumption_swmm_ids else "MODELED_FROM_SOURCE"
        else:
            if swmm_id in dropped_swmm_ids:
                modeled_reason = "NOT_MODELED_DUPLICATE_CONFLICT"
            elif row["structure_type"] == "inlet":
                basin = next((b for b in basins if b["index"] == row["basin_index"]), None)
                recv = None
                if basin:
                    recv = next((jr for jr in reversed(basin.get("junction_rows", [])) if jr["row_idx"] < row["source_row"]), None)
                modeled_reason = "NOT_MODELED_MISSING_RECEIVING_JUNCTION" if recv is None else "NOT_MODELED_DUPLICATE_CONFLICT"
            else:
                modeled_reason = "NOT_MODELED_DUPLICATE_CONFLICT"
        unaccounted_rows.append({**row, "modeled": modeled, "modeled_reason": modeled_reason, "assumptions_used": swmm_id in assumption_swmm_ids})

    unaccounted_rows = []
    for rec in all_structures:
        sid = rec["jct_id"] if rec["structure_type"] == "junction" else rec["inlet_id"]
        swmm_id = j_name(sid) if rec["structure_type"] == "junction" else in_name(sid)
        unaccounted_rows.append({
            "source_tab": "HYDROLOGY",
            "source_row": rec["row_idx"],
            "basin_index": rec["basin_index"],
            "structure_type": rec["structure_type"],
            "source_id": sid,
            "swmm_id": swmm_id,
            "modeled": swmm_id in modeled_nodes,
        })
    pd.DataFrame(unaccounted_rows).to_csv(ROOT / "outputs/review/excel_unaccounted_structures.csv", index=False)

    outfall_links = {c["cid"] for c in conduits if c.get("link_type") == "outfall"}
    outfall_recon = []
    for basin in basins:
        if not basin.get("junction_rows"):
            continue
        term = basin["junction_rows"][-1]
        tid = j_name(term["jct_id"])
        oid = out_name(term["jct_id"])
        outfall_recon.append({
            "basin_index": basin["index"],
            "terminal_source_row": term["row_idx"],
            "terminal_junction": tid,
            "outfall_id": oid,
            "outfall_node_modeled": oid in outfalls,
            "outfall_conduit_modeled": f"L_{tid}__{oid}" in outfall_links,
        })
    pd.DataFrame(outfall_recon).to_csv(ROOT / "outputs/review/terminal_basin_outfall_reconciliation.csv", index=False)
    (ROOT / "outputs/review/phase2_terminal_branch_confusion.md").write_text(
        "# Phase 2 Terminal Branch Confusion\n\n"
        "Topology uses HYDROLOGY BASIN markers only. Each basin terminal junction discharges to basin-specific outfall O_<terminal_id>.\n\n"
        f"- Basins parsed: {len([b for b in basins if b.get('junction_rows')])}\n"
        f"- Outfall nodes modeled: {len(outfalls)}\n"
        f"- Outfall conduits modeled: {len(outfall_links)}\n",
        encoding="utf-8",
    )

    modeled_nodes = set(nodes.keys())
    unaccounted_rows = []
    for rec in all_structures:
        sid = rec["jct_id"] if rec["structure_type"] == "junction" else rec["inlet_id"]
        swmm_id = j_name(sid) if rec["structure_type"] == "junction" else in_name(sid)
        unaccounted_rows.append({
            "source_tab": "HYDROLOGY",
            "source_row": rec["row_idx"],
            "basin_index": rec["basin_index"],
            "structure_type": rec["structure_type"],
            "source_id": sid,
            "swmm_id": swmm_id,
            "modeled": swmm_id in modeled_nodes,
        })
    pd.DataFrame(unaccounted_rows).to_csv(ROOT / "outputs/review/excel_unaccounted_structures.csv", index=False)

    outfall_links = {c["cid"] for c in conduits if c.get("link_type") == "outfall"}
    outfall_rows = []
    for basin in basins:
        if not basin.get("junction_rows"):
            continue
        term = basin["junction_rows"][-1]
        tid = j_name(term["jct_id"])
        oid = out_name(term["jct_id"])
        outfall_rows.append({
            "basin_index": basin["index"],
            "terminal_source_row": term["row_idx"],
            "terminal_junction": tid,
            "outfall_id": oid,
            "outfall_node_modeled": oid in outfalls,
            "outfall_conduit_modeled": f"L_{tid}__{oid}" in outfall_links,
        })
    pd.DataFrame(outfall_rows).to_csv(ROOT / "outputs/review/terminal_basin_outfall_reconciliation.csv", index=False)
    (ROOT / "outputs/review/phase2_terminal_branch_confusion.md").write_text(
        "# Phase 2 Terminal Branch Confusion\n\n"
        "Topology uses HYDROLOGY BASIN markers only. Each basin terminal junction discharges to basin-specific outfall O_<terminal_id>.\n\n"
        f"- Basins parsed: {len([b for b in basins if b.get('junction_rows')])}\n"
        f"- Outfall nodes modeled: {len(outfalls)}\n"
        f"- Outfall conduits modeled: {len(outfall_links)}\n",
        encoding="utf-8",
    )

    junction_rows = sorted(nodes.items(), key=lambda kv: kv[0])
    outfall_rows = sorted(outfalls.items(), key=lambda kv: kv[0])
    inflow_rows = sorted([(k, v) for k, v in inflow_by_node.items() if v > 0], key=lambda t: t[0])

    text: list[str] = []
    text.append("[TITLE]\n;; Inlet-known-flow hydraulic routing baseline")
    text.append(
        "\n[OPTIONS]\n"
        "FLOW_UNITS           CFS\n"
        "INFILTRATION         HORTON\n"
        "FLOW_ROUTING         DYNWAVE\n"
        "START_DATE           01/01/2025\n"
        "START_TIME           00:00:00\n"
        "REPORT_START_DATE    01/01/2025\n"
        "REPORT_START_TIME    00:00:00\n"
        "END_DATE             01/01/2025\n"
        "END_TIME             06:00:00\n"
        "WET_STEP             00:05:00\n"
        "DRY_STEP             01:00:00\n"
        "ROUTING_STEP         00:00:30\n"
        "ALLOW_PONDING        NO"
    )
    text.append("\n[JUNCTIONS]\n;;Name  Elevation  MaxDepth  InitDepth  SurDepth  Aponded")
    for nid, d in junction_rows:
        text.append(f"{nid}  {d['elev']:.3f}  {d['max_depth']:.3f}  0.000  0.000  0.000")
    text.append("\n[OUTFALLS]\n;;Name  Elevation  Type  Stage Data  Gated  Route To")
    for oid, d in outfall_rows:
        text.append(f"{oid}  {d['elev']:.3f}  FREE")
    text.append("\n[CONDUITS]\n;;Name  FromNode  ToNode  Length  Roughness  InOffset  OutOffset  InitFlow  MaxFlow")
    for c in conduits:
        text.append(f"{c['cid']}  {c['us']}  {c['ds']}  {c['length']:.2f}  {roughness:.3f}  0.00  0.00  0.00  0.00")
    text.append("\n[XSECTIONS]\n;;Link  Shape  Geom1  Geom2  Geom3  Geom4  Barrels")
    for c in conduits:
        text.append(f"{c['cid']}  CIRCULAR  {c['dia_ft']:.3f}  0.000  0.000  0.000  1")
    text.append("\n[LOSSES]\n;;Link  Kentry  Kexit  Kavg  FlapGate  Seepage")
    for c in conduits:
        text.append(f"{c['cid']}  0.0  0.0  0.0  NO  0.0")
    text.append("\n[INFLOWS]\n;;Node  Constituent  TimeSeries  Type  Mfactor  Sfactor  Baseline  Pattern")
    for n, q in inflow_rows:
        text.append(f'{n}  FLOW  ""  FLOW  1.0  1.0  {q:.4f}  ""')
    text.append("\n[COORDINATES]\n;;Node  X-Coord  Y-Coord")
    for nid, _ in junction_rows:
        x, y = coords.get(nid, (0.0, 0.0))
        text.append(f"{nid}  {x:.2f}  {y:.2f}")
    for oid, _ in outfall_rows:
        x, y = coords.get(oid, (0.0, 0.0))
        text.append(f"{oid}  {x:.2f}  {y:.2f}")

    (ROOT / "outputs/qa/inlet_knownflow_hydraulic_baseline_qaqc_summary.md").write_text(
        "\n".join([
            "# Inlet Known-Flow Hydraulic Baseline QA/QC Summary",
            "",
            "- Phase scope: hydraulic routing only (no subcatchments, no rational runoff).",
            "- Inflows sourced from HYDROLOGY Column R and applied only at inlet nodes.",
            "- Inlet rim elevation assumption: inlet rim = receiving junction rim (HYDROLOGY AA).",
            f"- Junction nodes: {len(junction_rows)}",
            f"- Outfalls: {len(outfall_rows)}",
            f"- Conduits: {len(conduits)}",
            f"- Inflow nodes: {len(inflow_rows)}",
            f"- QA findings (HIGH): {sum(1 for f in findings if f['severity']=='HIGH')}",
            f"- QA findings (MEDIUM/LOW): {sum(1 for f in findings if f['severity']!='HIGH')}",
        ]) + "\n", encoding="utf-8")

    metadata = {
        "junction_count": len(junction_rows),
        "outfall_count": len(outfall_rows),
        "conduit_count": len(conduits),
        "subcatchment_count": 0,
        "inflow_count": len(inflow_rows),
        "assumptions": {
            "topology_mode": "inlet_aware_hydraulic_baseline",
            "outfall_mode": "basin_segmented",
            "subcatchments_included": False,
            "inflow_mode": "static_known_flows_at_inlets",
            "inlet_nodes_modeled_as": "junctions",
            "coordinates_mode": "orthogonal_schematic_from_lengths",
            "column_d_used_as_inlet_id_count": 0,
            "assumptions_enabled": assumptions_enabled,
            "min_slope_assumed": min_slope,
            "interpolation_enabled": assumptions_enabled,
        },
        "geometry_mode": "orthogonal_schematic_from_lengths",
        "assumptions_enabled": assumptions_enabled,
        "min_slope_assumed": min_slope,
        "interpolation_enabled": assumptions_enabled,
    }
    return "\n".join(text) + "\n", metadata


def run_build(ctx: PipelineContext) -> StageResult:
    processed = ROOT / "data/processed"
    if not (processed / "nodes.csv").exists() or not (processed / "links.csv").exists():
        return StageResult("build", "blocked", "Missing processed nodes/links", [])

    model_text, metadata = _build_runnable_model(ctx)
    model_path = ROOT / "models/model.inp"
    model_path.write_text(model_text, encoding="utf-8")

    meta_path = ROOT / "outputs/logs/build_summary.json"
    meta_path.write_text(json.dumps({"generated_at_utc": now_utc(), **metadata}, indent=2), encoding="utf-8")

    return StageResult("build", "ready", "Runnable SWMM model written", [str(model_path.relative_to(ROOT)), str(meta_path.relative_to(ROOT))])


def run_qa(ctx: PipelineContext) -> StageResult:
    import pandas as pd

    findings: List[dict] = []
    nodes_path = ROOT / "data/processed/nodes.csv"
    links_path = ROOT / "data/processed/links.csv"
    model_path = ROOT / "models/model.inp"

    if not nodes_path.exists() or not links_path.exists():
        findings.append({"severity": "high", "check": "processed_files", "message": "Missing processed nodes or links"})
    else:
        nodes_df = pd.read_csv(nodes_path) if nodes_path.stat().st_size > 0 else pd.DataFrame()
        links_df = pd.read_csv(links_path) if links_path.stat().st_size > 0 else pd.DataFrame()
        if nodes_df.empty:
            findings.append({"severity": "high", "check": "nodes_non_empty", "message": "nodes.csv is empty"})
        if links_df.empty:
            findings.append({"severity": "medium", "check": "links_non_empty", "message": "links.csv is empty; build uses synthetic chaining"})

    required_sections = [
        "[OPTIONS]",
        "[RAINGAGES]",
        "[TIMESERIES]",
        "[JUNCTIONS]",
        "[OUTFALLS]",
        "[CONDUITS]",
        "[XSECTIONS]",
        "[LOSSES]",
        "[SUBCATCHMENTS]",
        "[SUBAREAS]",
        "[INFILTRATION]",
        "[INFLOWS]",
        "[COORDINATES]",
    ]

    if not model_path.exists():
        findings.append({"severity": "high", "check": "model_exists", "message": "models/model.inp is missing"})
    else:
        model_text = model_path.read_text(encoding="utf-8")
        for section in required_sections:
            if section not in model_text:
                findings.append({"severity": "high", "check": "required_section", "message": f"Missing {section} in model.inp"})

    status = "ready" if not any(f["severity"] == "high" for f in findings) else "blocked"

    qa_dir = ROOT / "outputs/qa"
    qa_dir.mkdir(parents=True, exist_ok=True)
    findings_path = qa_dir / "qa_findings.csv"
    write_csv(findings_path, findings)

    summary = {
        "generated_at_utc": now_utc(),
        "status": status,
        "finding_count": len(findings),
        "high_count": sum(1 for f in findings if f.get("severity") == "high"),
    }
    summary_path = qa_dir / "qa_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    return StageResult("qa", status, "QA complete", [str(findings_path.relative_to(ROOT)), str(summary_path.relative_to(ROOT))])


def run_stage(stage: str) -> StageResult:
    ctx = PipelineContext()
    if stage == "preflight":
        return run_preflight(ctx)
    if stage == "extract":
        return run_extract(ctx)
    if stage == "transform":
        return run_transform(ctx)
    if stage == "build":
        return run_build(ctx)
    if stage == "qa":
        return run_qa(ctx)
    raise ValueError(f"Unknown stage: {stage}")
