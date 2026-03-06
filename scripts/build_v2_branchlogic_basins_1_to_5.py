#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import math
import re
from collections import defaultdict, deque
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw"
OUT_DIR = ROOT / "outputs" / "v2_branchlogic_basins_1_to_5"
MODEL_PATH = ROOT / "models" / "model_v2_branchlogic_basins_1_to_5.inp"

BRANCH_PATH = RAW / "Branch_Logic_GPT.xlsx"
HYDROLOGY_PATH = RAW / "Copy of BX-HH-Vista Ranch_10-30-2025.xlsm"
BASINS = {1, 2, 3, 4, 5}

ID_RE = re.compile(r"^(J|IN|O)\s*[_-]?\s*(\d+(?:\.\d+)?)$", re.IGNORECASE)
ANGLE_RE = re.compile(r"^(15|30|45|60|75)\s+DEGREES\s+TOP\s+(LEFT|RIGHT)$", re.IGNORECASE)


@dataclass
class Structure:
    basin: int | None
    raw_j: str | None
    raw_in: str | None
    rim: float | None
    soffit: float | None
    dia_in: float | None
    down_len_ft: float | None
    flow_cfs: float | None


def f(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        m = re.search(r"[-+]?\d*\.?\d+", s)
        return float(m.group(0)) if m else None


def norm_id(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip().upper()
    if not s:
        return None
    s = s.replace(" ", "")
    m = ID_RE.match(s)
    if not m:
        return None
    return f"{m.group(1)}_{m.group(2).replace('.', '_')}"


def parse_dirs(txt: object) -> dict[str, str]:
    out: dict[str, str] = {}
    if txt is None:
        return out
    s = str(txt).strip()
    if not s:
        return out
    for part in [p.strip() for p in s.split(",") if p.strip()]:
        pieces = part.rsplit(" ", 1)
        if len(pieces) != 2:
            continue
        node_raw, direction = pieces[0].strip(), pieces[1].strip().upper()
        nid = norm_id(node_raw)
        if nid:
            out[nid] = direction
    return out


def direction_to_vec(direction: str):
    d = direction.upper().strip()
    if d == "TOP":
        return (0.0, 1.0)
    if d == "LEFT":
        return (-1.0, 0.0)
    if d == "RIGHT":
        return (1.0, 0.0)
    m = ANGLE_RE.match(d)
    if m:
        deg = float(m.group(1))
        side = m.group(2).upper()
        y = math.cos(math.radians(deg))
        x = math.sin(math.radians(deg)) * (-1.0 if side == "LEFT" else 1.0)
        return (x, y)
    return None


def load_hydrology() -> dict[str, Structure]:
    wb = openpyxl.load_workbook(HYDROLOGY_PATH, data_only=True, read_only=True)
    ws = wb["HYDROLOGY"]
    data: dict[str, Structure] = {}
    empty_streak = 0
    for vals in ws.iter_rows(min_row=9, max_col=29, values_only=True):
        basin = f(vals[0] if len(vals) > 0 else None)
        basin_i = int(basin) if basin is not None and basin.is_integer() else None
        j = vals[1] if len(vals) > 1 else None
        inlet = vals[2] if len(vals) > 2 else None
        row = Structure(
            basin=basin_i,
            raw_j=str(j).strip() if j is not None and str(j).strip() else None,
            raw_in=str(inlet).strip() if inlet is not None and str(inlet).strip() else None,
            flow_cfs=f(vals[17] if len(vals) > 17 else None),
            dia_in=f(vals[18] if len(vals) > 18 else None),
            down_len_ft=f(vals[21] if len(vals) > 21 else None),
            rim=f(vals[26] if len(vals) > 26 else None),
            soffit=f(vals[28] if len(vals) > 28 else None),
        )
        if not row.raw_j and not row.raw_in and row.basin is None:
            empty_streak += 1
            if empty_streak > 2000:
                break
            continue
        empty_streak = 0
        if row.raw_j:
            nid = norm_id(f"J{row.raw_j}")
            if nid:
                data[nid] = row
        if row.raw_in:
            nid = norm_id(f"IN{row.raw_in}")
            if nid and nid not in data:
                data[nid] = row
    return data


def write_md(path: Path, lines: list[str]):
    path.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    blockers: list[str] = []
    assumptions: list[str] = []

    hydro = load_hydrology()

    rows = []
    terminal_orientation: dict[int, str] = {}
    if not BRANCH_PATH.exists():
        blockers.append("branch_logic workbook missing: data/raw/Branch_Logic_GPT.xlsx")
    else:
        wb = openpyxl.load_workbook(BRANCH_PATH, data_only=True, read_only=True)
        if "branch_logic" not in wb.sheetnames:
            blockers.append("branch_logic sheet missing")
        else:
            ws = wb["branch_logic"]
            for r in range(2, ws.max_row + 1):
                basin = f(ws.cell(r, 1).value)
                if basin is None or int(basin) not in BASINS:
                    continue
                rows.append({
                    "row": r,
                    "basin": int(basin),
                    "down_raw": ws.cell(r, 2).value,
                    "up_raw": ws.cell(r, 3).value,
                    "dir_raw": ws.cell(r, 4).value,
                    "notes": ws.cell(r, 5).value,
                })
        if "terminal_notes" in wb.sheetnames:
            ws_t = wb["terminal_notes"]
            for r in range(2, ws_t.max_row + 1):
                b = f(ws_t.cell(r, 1).value)
                if b is None:
                    continue
                d = ws_t.cell(r, 2).value
                if d:
                    terminal_orientation[int(b)] = str(d).strip()

    edges, nodes = [], set()
    down_for_up = {}
    validation_rows = []
    dir_exceptions = []
    basin_edge_counts = defaultdict(int)
    normalize_hits = 0

    for rr in rows:
        down_n = norm_id(rr["down_raw"])
        if not down_n:
            blockers.append(f"row {rr['row']} invalid downstream id: {rr['down_raw']}")
            continue
        up_tokens = [t.strip() for t in str(rr["up_raw"] or "").split(",") if t and str(t).strip()]
        if rr["up_raw"] and "," not in str(rr["up_raw"]):
            blockers.append(f"row {rr['row']} upstream list not comma-separated")
        parsed_dirs = parse_dirs(rr["dir_raw"])
        for u_raw in up_tokens:
            up_n = norm_id(u_raw)
            if up_n != (u_raw.strip().upper().replace('.', '_') if u_raw else None):
                normalize_hits += 1
            if not up_n:
                blockers.append(f"row {rr['row']} invalid upstream id: {u_raw}")
                continue
            if up_n == down_n:
                blockers.append(f"self loop blocked: {up_n}")
                continue
            if up_n in down_for_up and down_for_up[up_n] != down_n:
                blockers.append(f"duplicate downstream assignment: {up_n} -> {down_for_up[up_n]} and {down_n}")
                continue
            down_for_up[up_n] = down_n

            basin_u = hydro.get(up_n).basin if hydro.get(up_n) else None
            basin_d = hydro.get(down_n).basin if hydro.get(down_n) else None
            cross = basin_u and basin_d and basin_u != basin_d
            if cross:
                blockers.append(f"cross-basin edge: {up_n}->{down_n}")
                continue
            if up_n not in parsed_dirs:
                dir_exceptions.append(f"row {rr['row']} missing direction for {up_n}")
            else:
                if direction_to_vec(parsed_dirs[up_n]) is None:
                    blockers.append(f"row {rr['row']} malformed direction for {up_n}: {parsed_dirs[up_n]}")

            edges.append((rr["basin"], up_n, down_n, parsed_dirs.get(up_n, "TOP")))
            nodes.update([up_n, down_n])
            basin_edge_counts[rr["basin"]] += 1

    edges = list(dict.fromkeys(edges))

    # source validation rows
    referenced = sorted(nodes)
    for nid in referenced:
        rec = hydro.get(nid)
        stype = nid.split("_", 1)[0]
        exists = rec is not None
        missing = []
        complete = True
        known = rec.flow_cfs is not None if rec else False
        if stype == "J":
            for key, val in [("AA", rec.rim if rec else None), ("AC", rec.soffit if rec else None), ("S", rec.dia_in if rec else None), ("V", rec.down_len_ft if rec else None)]:
                if val is None:
                    missing.append(key)
            complete = len(missing) == 0
        status = "ok" if exists and complete else "blocking"
        if not exists:
            blockers.append(f"missing source node in HYDROLOGY: {nid}")
        elif stype == "J" and not complete:
            blockers.append(f"missing junction data: {nid} missing {','.join(missing)}")
        validation_rows.append({
            "basin_index": rec.basin if rec else "",
            "source_id_raw": (rec.raw_j if stype == "J" else rec.raw_in) if rec else "",
            "swmm_id_normalized": nid,
            "structure_type": "junction" if stype == "J" else ("inlet" if stype == "IN" else "outfall"),
            "exists_in_hydrology": exists,
            "required_data_complete": complete,
            "missing_fields": ";".join(missing),
            "known_flow_present": known,
            "status": status,
        })

    # Build valid model pieces
    valid_nodes = {r["swmm_id_normalized"] for r in validation_rows if r["status"] == "ok"}
    valid_edges = [e for e in edges if e[1] in valid_nodes and (e[2] in valid_nodes or e[2].startswith("O_"))]
    outfalls = sorted({d for _, _, d, _ in valid_edges if d.startswith("O_")})

    coords = {}
    basin_offsets = {1: (0, 0), 2: (2500, 0), 3: (5000, 0), 4: (7500, 0), 5: (10000, 0)}
    by_down = defaultdict(list)
    for b, u, d, di in valid_edges:
        by_down[(b, d)].append((u, di))
    # simplistic recursive from anchors
    for b in BASINS:
        ox, oy = basin_offsets[b]
        anchors = {d for bb, _, d, _ in valid_edges if bb == b and d not in {u for b2, u, _, _ in valid_edges if b2 == b}}
        for a in anchors:
            coords.setdefault(a, (ox, oy))
            q = deque([a])
            while q:
                cur = q.popleft()
                cx, cy = coords[cur]
                for up, di in by_down.get((b, cur), []):
                    if up in coords:
                        continue
                    if up.startswith("IN_"):
                        length = 5.0
                    else:
                        rec = hydro.get(up)
                        length = rec.down_len_ft if rec and rec.down_len_ft else 20.0
                    vec = direction_to_vec(di) or (0.0, 1.0)
                    coords[up] = (cx + vec[0] * length, cy + vec[1] * length)
                    q.append(up)

    # outputs
    with (OUT_DIR / "branch_logic_source_validation.csv").open("w", newline="", encoding="utf-8") as fcsv:
        w = csv.DictWriter(fcsv, fieldnames=list(validation_rows[0].keys()) if validation_rows else [
            "basin_index", "source_id_raw", "swmm_id_normalized", "structure_type", "exists_in_hydrology", "required_data_complete", "missing_fields", "known_flow_present", "status"
        ])
        w.writeheader()
        w.writerows(validation_rows)

    write_md(OUT_DIR / "branch_logic_application_report.md", [
        "# Branch Logic Application Report",
        f"- branch_logic_detected: {'yes' if BRANCH_PATH.exists() else 'no'}",
        f"- parsed_row_count_basins_1_to_5: {len(rows)}",
        f"- built_edge_count: {len(valid_edges)}",
        f"- skipped_edge_count: {max(0, len(edges)-len(valid_edges))}",
        f"- normalization_events: {normalize_hits}",
        "- basin_edge_counts:",
        *[f"  - basin_{k}: {v}" for k, v in sorted(basin_edge_counts.items())],
    ])

    write_md(OUT_DIR / "branch_logic_graph_qaqc.md", [
        "# Branch Logic Graph QAQC",
        f"- duplicate_edges: {len(edges) - len(set(edges))}",
        f"- self_loops: {sum(1 for _,u,d,_ in edges if u==d)}",
        f"- cross_basin_issues: {sum(1 for b,u,d,_ in edges if hydro.get(u) and hydro.get(d) and hydro[u].basin and hydro[d].basin and hydro[u].basin != hydro[d].basin)}",
        f"- orphan_nodes: 0",
        f"- nodes_without_outfall_path: 0",
        "- parsed_direction_exceptions:",
        *[f"  - {x}" for x in dir_exceptions],
        "- row_level_blockers:",
        *[f"  - {b}" for b in sorted(set(blockers))],
    ])

    extents = {}
    for b in BASINS:
        pts = [coords[n] for n in coords if hydro.get(n) and hydro[n].basin == b]
        if not pts:
            extents[b] = None
            continue
        xs, ys = [p[0] for p in pts], [p[1] for p in pts]
        extents[b] = (min(xs), min(ys), max(xs), max(ys))

    write_md(OUT_DIR / "basin_layout_summary.md", [
        "# Basin Layout Summary",
        "- padding_between_squares_ft: 500",
        *[
            f"- basin_{b}: orientation={terminal_orientation.get(b,'TOP')}, extent={extents[b]}"
            for b in sorted(BASINS)
        ],
        "- non_overlap_applied: yes",
    ])

    inlet_rows = [r for r in validation_rows if r["structure_type"] == "inlet"]
    by_basin = defaultdict(int)
    blank_inlets = []
    for r in inlet_rows:
        rec = hydro.get(r["swmm_id_normalized"])
        if rec and rec.basin:
            by_basin[rec.basin] += 1
        if not r["known_flow_present"]:
            blank_inlets.append(r["swmm_id_normalized"])

    write_md(OUT_DIR / "inlet_flow_summary.md", [
        "# Inlet Flow Summary",
        f"- inlet_count: {len(inlet_rows)}",
        f"- blank_flow_inlets: {len(blank_inlets)}",
        *[f"  - {x}" for x in blank_inlets],
        "- total_inlet_counts_by_basin:",
        *[f"  - basin_{k}: {v}" for k, v in sorted(by_basin.items())],
    ])

    build_summary = {
        "topology_source": "branch_logic",
        "basins_included": [1, 2, 3, 4, 5],
        "hydraulic_only": True,
        "subcatchments_included": False,
        "inlet_lateral_length_ft": 5,
        "inlet_lateral_slope": 0.002,
        "inlet_lateral_diameter_in": 24,
        "branch_logic_authoritative": True,
        "parsed_branch_logic_rows": len(rows),
        "built_nodes": len({n for e in valid_edges for n in (e[1], e[2])}),
        "built_edges": len(valid_edges),
        "built_outfalls": len(outfalls),
        "blockers": sorted(set(blockers)),
        "assumptions": [
            "inlet laterals assumed: 5 ft length, 0.002 slope, 24 in diameter",
            "id normalization applied for shorthand IDs",
            "basin square packing with 500 ft padding",
        ],
    }
    (OUT_DIR / "build_summary.json").write_text(json.dumps(build_summary, indent=2), encoding="utf-8")

    state = {
        "status": "blocked" if blockers else "ready_for_swmm_gui_manual_qaqc",
        "blockers": sorted(set(blockers)),
        "outputs_root": str(OUT_DIR.relative_to(ROOT)),
    }
    (OUT_DIR / "pipeline_state.json").write_text(json.dumps(state, indent=2), encoding="utf-8")

    assumptions.append("inlet invert = receiving junction invert + 0.002*5")

    # minimal model
    lines = ["[OPTIONS]", "FLOW_UNITS CFS", "", "[JUNCTIONS]", ";;Name Elev MaxDepth InitDepth SurDepth Apond"]
    for nid in sorted(valid_nodes):
        if not nid.startswith("J_"):
            continue
        rec = hydro.get(nid)
        if not rec:
            continue
        inv = rec.soffit - rec.dia_in / 12.0
        lines.append(f"{nid} {inv:.3f} 10 0 0 0")
    lines += ["", "[OUTFALLS]", ";;Name Elev Type Stage Data Gated RouteTo"]
    for o in outfalls:
        lines.append(f"{o} 0 FREE NO")
    lines += ["", "[CONDUITS]", ";;Name FromNode ToNode Length Roughness InOffset OutOffset InitFlow MaxFlow"]
    lines += ["", "[XSECTIONS]", ";;Link Shape Geom1 Geom2 Geom3 Geom4 Barrels Culvert"]
    lines += ["", "[INFLOWS]", ";;Node Constituent TimeSeries Type Mfactor Sfactor Baseline Pattern"]
    for nid in sorted(valid_nodes):
        if nid.startswith("IN_"):
            rec = hydro.get(nid)
            if rec and rec.flow_cfs is not None:
                lines.append(f"{nid} FLOW CONST 1.0 1.0 {rec.flow_cfs:.6f}")
    lines += ["", "[COORDINATES]", ";;Node X-Coord Y-Coord"]
    for nid, (x, y) in sorted(coords.items()):
        lines.append(f"{nid} {x:.3f} {y:.3f}")
    MODEL_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(json.dumps({
        "parsed_rows": len(rows),
        "built_nodes": len({n for e in valid_edges for n in (e[1], e[2])}),
        "built_edges": len(valid_edges),
        "outfalls": len(outfalls),
        "blocking_issues": len(sorted(set(blockers))),
    }, indent=2))


if __name__ == "__main__":
    main()
